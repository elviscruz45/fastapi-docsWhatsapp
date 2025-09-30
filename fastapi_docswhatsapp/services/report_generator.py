from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Tuple
import os

from fastapi_docswhatsapp.models import ProjectAnalysis

class ReportGenerator:
    """Clase para generar reportes en PDF y Excel"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Configura estilos personalizados para el PDF"""
        # Estilo para títulos principales
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#2E86AB'),
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        
        # Estilo para subtítulos
        self.styles.add(ParagraphStyle(
            name='CustomHeading2',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#A23B72'),
            spaceBefore=20,
            spaceAfter=10
        ))
        
        # Estilo para texto normal
        self.styles.add(ParagraphStyle(
            name='CustomNormal',
            parent=self.styles['Normal'],
            fontSize=11,
            textColor=colors.black,
            spaceAfter=6,
            alignment=TA_JUSTIFY
        ))
    
    async def generate_reports(self, chat_data: Dict[str, Any], analysis: ProjectAnalysis, 
                              output_dir: Path) -> Tuple[Path, Path]:
        """
        Genera tanto el reporte PDF como Excel
        """
        # Generar nombres de archivos únicos
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        chat_name = chat_data.get('chat_name', 'chat').replace(' ', '_')
        
        pdf_path = output_dir / f"reporte_{chat_name}_{timestamp}.pdf"
        excel_path = output_dir / f"datos_{chat_name}_{timestamp}.xlsx"
        
        # Generar ambos reportes
        await self.generate_pdf_report(chat_data, analysis, pdf_path)
        await self.generate_excel_report(chat_data, analysis, excel_path)
        
        return pdf_path, excel_path
    
    async def generate_pdf_report(self, chat_data: Dict[str, Any], analysis: ProjectAnalysis, 
                                  output_path: Path):
        """Genera el reporte en formato PDF"""
        doc = SimpleDocTemplate(str(output_path), pagesize=A4)
        story = []
        
        # Título del reporte
        title = f"Reporte de Análisis de Proyecto\n{chat_data.get('chat_name', 'Chat de WhatsApp')}"
        story.append(Paragraph(title, self.styles['CustomTitle']))
        story.append(Spacer(1, 20))
        
        # Información general
        story.append(Paragraph("📋 Información General", self.styles['CustomHeading2']))
        
        general_info = [
            ['Nombre del Chat:', chat_data.get('chat_name', 'N/A')],
            ['Participantes:', ', '.join(chat_data.get('participants', []))],
            ['Total de Mensajes:', str(chat_data.get('total_messages', 0))],
            ['Período:', f"{chat_data.get('date_range', {}).get('start', 'N/A')} - {chat_data.get('date_range', {}).get('end', 'N/A')}"],
            ['Fecha de Análisis:', datetime.now().strftime("%d/%m/%Y %H:%M")]
        ]
        
        info_table = Table(general_info, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Resumen ejecutivo
        story.append(Paragraph("📊 Resumen Ejecutivo", self.styles['CustomHeading2']))
        story.append(Paragraph(analysis.summary, self.styles['CustomNormal']))
        story.append(Spacer(1, 15))
        
        # Hitos clave
        story.append(Paragraph("🎯 Hitos Clave Identificados", self.styles['CustomHeading2']))
        for i, milestone in enumerate(analysis.key_milestones, 1):
            story.append(Paragraph(f"{i}. {milestone}", self.styles['CustomNormal']))
        story.append(Spacer(1, 15))
        
        # Indicadores de progreso
        if analysis.progress_indicators:
            story.append(Paragraph("📈 Indicadores de Progreso", self.styles['CustomHeading2']))
            
            progress_data = [['Indicador', 'Valor', 'Descripción']]
            for indicator in analysis.progress_indicators:
                progress_data.append([
                    indicator.get('indicator', ''),
                    indicator.get('value', ''),
                    indicator.get('description', '')[:50] + '...' if len(indicator.get('description', '')) > 50 else indicator.get('description', '')
                ])
            
            progress_table = Table(progress_data, colWidths=[1.5*inch, 1*inch, 3.5*inch])
            progress_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')])
            ]))
            
            story.append(progress_table)
            story.append(Spacer(1, 15))
        
        # Nueva página para desafíos y recomendaciones
        story.append(PageBreak())
        
        # Desafíos identificados
        story.append(Paragraph("⚠️ Desafíos Identificados", self.styles['CustomHeading2']))
        for i, challenge in enumerate(analysis.challenges_identified, 1):
            story.append(Paragraph(f"{i}. {challenge}", self.styles['CustomNormal']))
        story.append(Spacer(1, 15))
        
        # Recomendaciones
        story.append(Paragraph("💡 Recomendaciones", self.styles['CustomHeading2']))
        for i, recommendation in enumerate(analysis.recommendations, 1):
            story.append(Paragraph(f"{i}. {recommendation}", self.styles['CustomNormal']))
        story.append(Spacer(1, 15))
        
        # Análisis de timeline
        if analysis.timeline_analysis:
            story.append(Paragraph("📅 Análisis de Cronograma", self.styles['CustomHeading2']))
            
            timeline_data = []
            timeline = analysis.timeline_analysis
            
            if timeline.get('project_start'):
                timeline_data.append(['Inicio del Proyecto:', timeline['project_start']])
            if timeline.get('current_phase'):
                timeline_data.append(['Fase Actual:', timeline['current_phase']])
            if timeline.get('estimated_completion'):
                timeline_data.append(['Finalización Estimada:', timeline['estimated_completion']])
            
            if timeline_data:
                timeline_table = Table(timeline_data, colWidths=[2*inch, 4*inch])
                timeline_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                
                story.append(timeline_table)
            story.append(Spacer(1, 15))
        
        # Contribuciones por participante
        if analysis.participant_contributions:
            story.append(Paragraph("👥 Contribuciones por Participante", self.styles['CustomHeading2']))
            
            for participant, contribution in analysis.participant_contributions.items():
                story.append(Paragraph(f"<b>{participant}:</b> {contribution}", self.styles['CustomNormal']))
            story.append(Spacer(1, 15))
        
        # Pie de página
        story.append(Spacer(1, 30))
        story.append(Paragraph(
            f"Reporte generado automáticamente el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}",
            self.styles['Normal']
        ))
        
        # Construir PDF
        doc.build(story)
    
    async def generate_excel_report(self, chat_data: Dict[str, Any], analysis: ProjectAnalysis,
                                   output_path: Path):
        """Genera el reporte en formato Excel"""
        wb = openpyxl.Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # Crear hojas
        self._create_summary_sheet(wb, chat_data, analysis)
        self._create_messages_sheet(wb, chat_data)
        self._create_analysis_sheet(wb, analysis)
        self._create_timeline_sheet(wb, analysis)
        
        # Guardar archivo
        wb.save(output_path)
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, chat_data: Dict[str, Any], 
                             analysis: ProjectAnalysis):
        """Crea la hoja de resumen"""
        ws = wb.create_sheet("Resumen", 0)
        
        # Configurar estilos
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        
        # Título
        ws['A1'] = "REPORTE DE ANÁLISIS DE PROYECTO WHATSAPP"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Información general
        ws['A3'] = "INFORMACIÓN GENERAL"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws.merge_cells('A3:B3')
        
        data = [
            ("Nombre del Chat:", chat_data.get('chat_name', 'N/A')),
            ("Total de Mensajes:", chat_data.get('total_messages', 0)),
            ("Participantes:", len(chat_data.get('participants', []))),
            ("Fecha de Análisis:", datetime.now().strftime("%d/%m/%Y %H:%M"))
        ]
        
        for i, (label, value) in enumerate(data, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Resumen ejecutivo
        ws['A9'] = "RESUMEN EJECUTIVO"
        ws['A9'].font = header_font
        ws['A9'].fill = header_fill
        ws.merge_cells('A9:D9')
        
        ws['A10'] = analysis.summary
        ws['A10'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('A10:D12')
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
    
    def _create_messages_sheet(self, wb: openpyxl.Workbook, chat_data: Dict[str, Any]):
        """Crea la hoja de mensajes"""
        ws = wb.create_sheet("Mensajes")
        
        # Encabezados
        headers = ["Fecha/Hora", "Remitente", "Tipo", "Contenido"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="A23B72", end_color="A23B72", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Datos de mensajes
        messages = chat_data.get('messages', [])
        for row, message in enumerate(messages[:1000], start=2):  # Limitar a 1000 mensajes
            ws.cell(row=row, column=1, value=message.timestamp.strftime("%d/%m/%Y %H:%M"))
            ws.cell(row=row, column=2, value=message.sender)
            ws.cell(row=row, column=3, value=message.message_type)
            ws.cell(row=row, column=4, value=message.content[:200])  # Limitar contenido
        
        # Ajustar columnas
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
    
    def _create_analysis_sheet(self, wb: openpyxl.Workbook, analysis: ProjectAnalysis):
        """Crea la hoja de análisis detallado"""
        ws = wb.create_sheet("Análisis Detallado")
        
        # Hitos
        ws['A1'] = "HITOS CLAVE"
        ws['A1'].font = Font(bold=True, size=12)
        
        for i, milestone in enumerate(analysis.key_milestones, start=2):
            ws[f'A{i}'] = f"{i-1}. {milestone}"
        
        # Desafíos
        start_row = len(analysis.key_milestones) + 4
        ws[f'A{start_row}'] = "DESAFÍOS IDENTIFICADOS"
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        
        for i, challenge in enumerate(analysis.challenges_identified, start=start_row+1):
            ws[f'A{i}'] = f"{i-start_row}. {challenge}"
        
        # Recomendaciones
        start_row = start_row + len(analysis.challenges_identified) + 3
        ws[f'A{start_row}'] = "RECOMENDACIONES"
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        
        for i, recommendation in enumerate(analysis.recommendations, start=start_row+1):
            ws[f'A{i}'] = f"{i-start_row}. {recommendation}"
        
        ws.column_dimensions['A'].width = 80
    
    def _create_timeline_sheet(self, wb: openpyxl.Workbook, analysis: ProjectAnalysis):
        """Crea la hoja de cronograma"""
        ws = wb.create_sheet("Cronograma")
        
        # Encabezados
        ws['A1'] = "ANÁLISIS DE CRONOGRAMA"
        ws['A1'].font = Font(bold=True, size=14)
        
        timeline = analysis.timeline_analysis
        
        data = [
            ("Inicio del Proyecto:", timeline.get('project_start', 'N/A')),
            ("Fase Actual:", timeline.get('current_phase', 'N/A')),
            ("Finalización Estimada:", timeline.get('estimated_completion', 'N/A'))
        ]
        
        for i, (label, value) in enumerate(data, start=3):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Fechas clave
        if timeline.get('key_dates'):
            ws['A7'] = "FECHAS CLAVE:"
            ws['A7'].font = Font(bold=True)
            
            for i, date_info in enumerate(timeline['key_dates'], start=8):
                ws[f'A{i}'] = date_info
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30